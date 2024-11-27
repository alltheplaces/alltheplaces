import io

from openpyxl import load_workbook
from scrapy import Spider

from locations.items import Feature


class YettelBGSpider(Spider):
    name = "yettel_bg"
    item_attributes = {
        "brand": "Yettel",
        "brand_wikidata": "Q14915070",
        "country": "BG",
    }
    start_urls = ["https://www.yettel.bg/faq/digital-customer-service/store-locator"]
    no_refs = True
    custom_settings = {"ROBOTSTXT_OBEY": False}

    def parse(self, response):
        yield response.follow(
            url=response.xpath('//input[@id="hdnExcelFile"]/@value').get(), callback=self.parse_spreadsheet
        )

    def parse_spreadsheet(self, response):
        if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get(
            "Content-Type"
        ).decode("utf-8"):
            excel_file = response.body

            excel_data = io.BytesIO(excel_file)
            workbook = load_workbook(excel_data, read_only=True)

            sheet = workbook.active

            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            headers = data[0]
            json_data = []
            for row in data[1:]:
                json_data.append({headers[i]: cell for i, cell in enumerate(row)})

            for store in json_data:
                item = Feature()

                item["lat"] = store["latitude"]
                item["lon"] = store["longitude"]

                item["street_address"] = store["address_loc"]
                item["city"] = store["city_loc"]

                yield item
