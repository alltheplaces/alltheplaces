import io

from openpyxl import load_workbook
from scrapy import Spider

from locations.categories import Categories, apply_category
from locations.items import Feature
from locations.pipelines.address_clean_up import clean_address

ZA_PROVINCES = {
    "EC": "Eastern Cape",
    "FS": "Free State",
    "GT": "Gauteng",
    "KZN": "KwaZulu-Natal",
    "LIM": "Limpopo",
    "MP": "Mpumalanga",
    "NC": "Northern Cape",
    "NW": "North West",
    "WC": "Western Cape",
}


class DbeGovEcdZASpider(Spider):
    name = "dbe_gov_ecd_za"
    download_timeout = 60

    # Link obtained from https://www.education.gov.za/Programmes/EMIS/EMISDownloads.aspx
    # It doesn't look like it can be reliably fetched if the page updates with newer data, so requires manually updating
    start_urls = [
        "https://www.education.gov.za/LinkClick.aspx?fileticket=c2vrRwdZJG8%3d&tabid=466&portalid=0&mid=10308"
    ]
    no_refs = True

    def parse(self, response):
        excel_file = response.body

        excel_data = io.BytesIO(excel_file)
        workbook = load_workbook(excel_data, read_only=True)

        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        headers = data[3]
        json_data = []
        for row in data[4:]:
            json_data.append({headers[i]: cell for i, cell in enumerate(row)})

        for location in json_data:
            item = Feature()

            apply_category(Categories.KINDERGARTEN, item)

            item["lat"] = location["GIS_Lat"]
            item["lon"] = location["GIS_Long"]

            if location.get("Town_City") is not None:
                item["city"] = str(location.get("Town_City")).title().replace("'S", "'s")

            item["state"] = ZA_PROVINCES.get(location["Province"])

            if location.get("StreetAddress") is not None:
                item["street_address"] = (
                    clean_address(str(location.get("StreetAddress"))).title().replace("'S", "'s").replace("`S", "'s")
                )

            if location.get("PostalAddress") is not None:
                item["extras"]["addr:postal"] = (
                    clean_address(str(location.get("PostalAddress"))).title().replace("'S", "'s").replace("`S", "'s")
                )

            item["name"] = str(location.get("ECD_Name")).title().replace("'S", "'s").replace("`S", "'s")

            item["phone"] = location["Telephone"]

            yield item
