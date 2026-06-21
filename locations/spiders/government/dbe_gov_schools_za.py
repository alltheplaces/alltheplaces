import io
import re
from typing import Any

from openpyxl import load_workbook
from scrapy import Spider
from scrapy.http import Response

from locations.categories import Categories, Extras, apply_category, apply_yes_no
from locations.items import Feature
from locations.pipelines.address_clean_up import clean_address

ZA_EDUCATION_DEPARTMENTS = {
    "EC": {"operator_wikidata": "Q116681237"},
    "FS": {"operator_wikidata": "Q116681212"},
    "GT": {"operator_wikidata": "Q116681267"},
    "KZN": {"operator_wikidata": "Q116681188"},
    "LP": {"operator_wikidata": "Q116681211"},
    "MP": {"operator_wikidata": "Q116681266"},
    "NC": {"operator_wikidata": "Q96396466"},
    "NW": {"operator_wikidata": "Q116681285"},
    "WC": {"operator_wikidata": "Q116681286"},
}

ZA_PROVINCES = {
    "EC": "Eastern Cape",
    "FS": "Free State",
    "GT": "Gauteng",
    "KZN": "KwaZulu-Natal",
    "LP": "Limpopo",
    "MP": "Mpumalanga",
    "NC": "Northern Cape",
    "NW": "North West",
    "WC": "Western Cape",
}


class DbeGovSchoolsZASpider(Spider):
    name = "dbe_gov_schools_za"
    custom_settings = {"DOWNLOAD_TIMEOUT": 60}

    # Links obtained from https://www.education.gov.za/Programmes/EMIS/EMISDownloads.aspx
    # It doesn't look like they can be reliably fetched if the page updates with newer data, so require manually updating
    start_urls = [
        "https://www.education.gov.za/LinkClick.aspx?fileticket=e6lMsXIcB7o%3d&tabid=466&portalid=0&mid=14616",  # Eastern Cape
        "https://www.education.gov.za/LinkClick.aspx?fileticket=PseEZWkRtV4%3d&tabid=466&portalid=0&mid=14616",  # Free State
        "https://www.education.gov.za/LinkClick.aspx?fileticket=mlHofRzcFX0%3d&tabid=466&portalid=0&mid=14616",  # Gauteng
        "https://www.education.gov.za/LinkClick.aspx?fileticket=Uo_ivbMxWpE%3d&tabid=466&portalid=0&mid=14616",  # KwaZulu-Natal
        "https://www.education.gov.za/LinkClick.aspx?fileticket=4vuFhKYNH_M%3d&tabid=466&portalid=0&mid=14616",  # Limpopo
        "https://www.education.gov.za/LinkClick.aspx?fileticket=6SXgSqRAckM%3d&tabid=466&portalid=0&mid=14616",  # Mpumalanga
        "https://www.education.gov.za/LinkClick.aspx?fileticket=1F4TEaSVt34%3d&tabid=466&portalid=0&mid=14616",  # North West
        "https://www.education.gov.za/LinkClick.aspx?fileticket=zPQtOgy47iE%3d&tabid=466&portalid=0&mid=14616",  # Northern Cape
        "https://www.education.gov.za/LinkClick.aspx?fileticket=gTqCBRz5VAI%3d&tabid=466&portalid=0&mid=14616",  # Western Cape
        "https://www.education.gov.za/LinkClick.aspx?fileticket=8x1LctBf_pw%3d&tabid=466&portalid=0&mid=14616",  # Special Needs Educatiom
    ]

    def parse(self, response: Response, **kwargs: Any) -> Any:
        excel_file = response.body

        excel_data = io.BytesIO(excel_file)
        workbook = load_workbook(excel_data, read_only=True)

        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        headers = data[0]
        json_data = []
        for row in data[1:]:
            json_data.append({headers[i]: cell for i, cell in enumerate(row)})

        for location in json_data:
            item = Feature()

            item["ref"] = location["NatEmis"]
            item["extras"]["ref:ZA:emis"] = location["NatEmis"]

            # Normalise column names
            if location["Province"] in ["GT"]:
                location["GIS_Latitude"] = location.get("GIS_Lat")
                location["GIS_Longitude"] = location.get("GIS_Long")

            item["lat"] = location.get("GIS_Latitude")
            item["lon"] = location.get("GIS_Longitude")

            # Coordinates are reversed in the data for most locations
            if (
                location["Province"] in ["EC"]
                and item.get("lon") is not None
                and item.get("lat") is not None
                and str(item["lon"])[0] == "-"
            ):
                item["lat"], item["lon"] = item["lon"], item["lat"]

            # Coordinates are stored without decimal point and reversed
            if location["Province"] in ["NC"] and item.get("lon") is not None and item.get("lat") is not None:
                item["lat"], item["lon"] = item["lon"], item["lat"]
                item["lat"] = str(item["lat"])[:3] + "." + str(item["lat"])[3:]
                item["lon"] = str(item["lon"])[:2] + "." + str(item["lon"])[2:]

            if location.get("Town_City") not in [None, 99]:
                item["city"] = location.get("Town_City").title().replace("'S", "'s")

            item["state"] = ZA_PROVINCES.get(location["Province"])

            if location.get("StreetAddress") not in [None, 99]:
                item["addr_full"] = (
                    clean_address(location.get("StreetAddress")).title().replace("'S", "'s").replace("`S", "'s")
                )

            if location.get("PostalAddress") not in [None, 99]:
                item["extras"]["addr:postal"] = (
                    clean_address(location.get("PostalAddress")).title().replace("'S", "'s").replace("`S", "'s")
                )

            if "addr_full" in item:
                if match := re.match(r"^.*\s(\d\d\d\d)$", item["addr_full"]):
                    item["postcode"] = match[1]

            item["name"] = (
                location.get("Official_Institution_Name")
                .title()
                .replace("'S", "'s")
                .replace("`S", "'s")
                .replace(" P/S", " Primary School")
                .replace(" S/S", " Secondary School")
                .replace(" C/S", " Combined School")
                .replace(" I/S", " Intermediate School")
            )

            item["phone"] = location["Telephone"]

            apply_yes_no(Extras.FEE, item, location["NoFeeSchool"] in ["Fee Charging"], True)

            if location["Sector"].lower() == "public":
                item.update(ZA_EDUCATION_DEPARTMENTS.get(location["Province"]))
                item["extras"]["operator:type"] = "government"
            elif location["Sector"].lower() == "independent":
                item["extras"]["operator:type"] = "private"
            else:
                self.crawler.stats.inc_value(f"atp/{self.name}/unknown_operator_tyoe/{location['Sector']}")

            self.set_category(item, location)

            yield item

    def set_category(self, item, location):
        if location["Phase_PED"] == "PRIMARY SCHOOL":
            item["extras"]["school"] = "primary"
            apply_category(Categories.SCHOOL, item)
        elif location["Phase_PED"] == "SECONDARY SCHOOL":
            item["extras"]["school"] = "secondary"
            apply_category(Categories.SCHOOL, item)
        elif location["Phase_PED"] == "COMBINED SCHOOL":
            item["extras"]["school"] = "primary;secondary"
            apply_category(Categories.SCHOOL, item)
        elif location["Phase_PED"] in ["SPECIAL NEEDS EDUCATION", "SPECIAL SCHOOL"]:
            item["extras"]["school"] = "special_education_needs"
            apply_category(Categories.SCHOOL, item)
        elif location["Phase_PED"] == "INTERMEDIATE SCHOOL":
            apply_category(Categories.SCHOOL, item)
        elif location["Phase_PED"] == "ECD":
            apply_category(Categories.KINDERGARTEN, item)
        elif (
            location["Phase_PED"] in ["HOSPITAL SCHOOL", "SCHOOL OF SKILLS"]
            or location["Phase_PED"] is None
            and location.get("Type_DoE") == "ORDINARY SCHOOL"
        ):
            apply_category(Categories.SCHOOL, item)
        else:
            self.crawler.stats.inc_value(f"atp/{self.name}/unknown_phase/{location['Phase_PED']}")

        if location["Type_DoE"] == "SPECIAL NEEDS EDUCATION CENTRE":
            item["extras"]["school"] = "special_education_needs"
            apply_category(Categories.SCHOOL, item)
