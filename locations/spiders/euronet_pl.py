import io
from typing import Any

from openpyxl import load_workbook
from scrapy import Spider
from scrapy.http import Response

from locations.categories import Categories, Extras, apply_category, apply_yes_no
from locations.dict_parser import DictParser
from locations.hours import DAYS, OpeningHours
from locations.spiders.alior_bank_pl import AliorBankPLSpider
from locations.spiders.millennium_bank_pl import MillenniumBankPLSpider
from locations.spiders.santander_pl import SantanderPLSpider
from locations.spiders.unicredit_bank_it import UnicreditBankITSpider

BRAND_MAPPING = {
    "Alior Bank": {
        "brand": AliorBankPLSpider.item_attributes["brand"],
        "brand_wikidata": AliorBankPLSpider.item_attributes["brand_wikidata"],
        "operator": AliorBankPLSpider.item_attributes["brand"],
        "operator_wikidata": AliorBankPLSpider.item_attributes["brand_wikidata"],
    },
    "Banku Millennium": {
        "brand": MillenniumBankPLSpider.item_attributes["brand"],
        "brand_wikidata": MillenniumBankPLSpider.item_attributes["brand_wikidata"],
        "operator": MillenniumBankPLSpider.item_attributes["brand"],
        "operator_wikidata": MillenniumBankPLSpider.item_attributes["brand_wikidata"],
    },
    "mBank": {
        "brand": "mBank",
        "brand_wikidata": "Q1160928",
        "operator": "mBank",
        "operator_wikidata": "Q1160928",
    },
    "mKiosk": {
        "brand": "mBank",
        "brand_wikidata": "Q1160928",
        "operator": "mBank",
        "operator_wikidata": "Q1160928",
    },
    "Santander Bank Polska": {
        "brand": SantanderPLSpider.item_attributes["brand"],
        "brand_wikidata": SantanderPLSpider.item_attributes["brand_wikidata"],
        "operator": SantanderPLSpider.item_attributes["brand"],
        "operator_wikidata": SantanderPLSpider.item_attributes["brand_wikidata"],
    },
    "Kasa Stefczyka": {
        "brand": "SKOK Stefczyka",
        "brand_wikidata": "Q57624461",
        "operator": "SKOK Stefczyka",
        "operator_wikidata": "Q57624461",
    },
    "UniCredit": {
        "brand": UnicreditBankITSpider.item_attributes["brand"],
        "brand_wikidata": UnicreditBankITSpider.item_attributes["brand_wikidata"],
        "operator": UnicreditBankITSpider.item_attributes["brand"],
        "operator_wikidata": UnicreditBankITSpider.item_attributes["brand_wikidata"],
    },
}


def parse_hours(hours_str: str) -> str | OpeningHours | None:
    """Parse hours like '00:00-23:59' into OpeningHours format"""
    if not hours_str or hours_str == "24/7":
        return "24/7"

    try:
        oh = OpeningHours()
        # Format appears to be HH:MM-HH:MM for all days
        if "-" in hours_str:
            start, end = hours_str.split("-")
            oh.add_days_range(DAYS, start.strip(), end.strip())
        return oh
    except:
        # If parsing fails, return None
        return None


class EuronetPLSpider(Spider):
    name = "euronet_pl"
    item_attributes = {"network": "Euronet", "network_wikidata": "Q5412010"}
    start_urls = ["https://euronet.pl/Lista_wplatomatow_sieci_Euronet.xlsx"]
    custom_settings = {"DOWNLOAD_TIMEOUT": 60}

    def parse(self, response: Response, **kwargs: Any) -> Any:
        excel_file = response.body
        excel_data = io.BytesIO(excel_file)
        workbook = load_workbook(excel_data, read_only=True)
        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        headers = data[0]
        json_data = []
        for row in data[1:]:
            json_data.append({headers[i]: cell for i, cell in enumerate(row)})

        for location in json_data:
            item = DictParser.parse(location)
            item["street_address"] = item.pop("addr_full")
            item["state"] = location.get("District")
            item["ref"] = location.get("ATM")

            for brand_key in BRAND_MAPPING:
                if brand_key in location["Name"]:
                    item.update(BRAND_MAPPING[brand_key])
                    break

            # Extract coordinates (replace comma with period for decimal separator)
            if lat := location.get("GPS_Latitude"):
                item["lat"] = str(lat).replace(",", ".")
            if lon := location.get("GPS_Longitude"):
                item["lon"] = str(lon).replace(",", ".")

            # Opening hours
            if hours := location.get("Client access hours"):
                item["opening_hours"] = parse_hours(hours)

            # Apply ATM category
            apply_category(Categories.ATM, item)
            apply_yes_no(Extras.CASH_IN, item, location.get("Typ") == "Recycler")

            yield item
