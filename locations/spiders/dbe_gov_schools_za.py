import io
import re

from openpyxl import load_workbook
from scrapy import Spider

from locations.categories import Categories, Extras, apply_category, apply_yes_no
from locations.items import Feature
from locations.pipelines.address_clean_up import clean_address

ZA_EDUCATION_DEPARTMENTS = {
    "EC": {"operator_wikidata": "Q116681237"},
    "FS": {"operator_wikidata": "Q116681212"},
    "GT": {"operator_wikidata": "Q116681267"},
    "KZN": {"operator_wikidata": "Q116681188"},
    "LP": {"operator_wikidata": "Q116681211"},
    "MP": {"operator_wikidata": "Q116681266"},
    "NC": {"operator_wikidata": "Q96396466"},
    "NW": {"operator_wikidata": "Q116681285"},
    "WC": {"operator_wikidata": "Q116681286"},
}

ZA_PROVINCES = {
    "EC": "Eastern Cape",
    "FS": "Free State",
    "GT": "Gauteng",
    "KZN": "KwaZulu-Natal",
    "LP": "Limpopo",
    "MP": "Mpumalanga",
    "NC": "Northern Cape",
    "NW": "North West",
    "WC": "Western Cape",
}


class DbeGovSchoolsZASpider(Spider):
    name = "dbe_gov_schools_za"
    download_timeout = 60

    # Links obtained from https://www.education.gov.za/Programmes/EMIS/EMISDownloads.aspx
    # It doesn't look like they can be reliably fetched if the page updates with newer data, so require manually updating
    start_urls = [
        "https://www.education.gov.za/LinkClick.aspx?fileticket=O2qbsfuGFC4%3d&tabid=466&portalid=0&mid=12484",  # National list
        "https://www.education.gov.za/LinkClick.aspx?fileticket=OtSySF85mAQ%3d&tabid=466&portalid=0&mid=12484",  # Special Needs Educatiom
    ]

    def parse(self, response):
        excel_file = response.body

        excel_data = io.BytesIO(excel_file)
        workbook = load_workbook(excel_data, read_only=True)

        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        headers = data[0]
        json_data = []
        for row in data[1:]:
            json_data.append({headers[i]: cell for i, cell in enumerate(row)})

        for location in json_data:
            item = Feature()

            item["ref"] = location["NatEmis"]
            item["extras"]["ref:ZA:emis"] = location["NatEmis"]

            item["lat"] = location["GIS_Latitude"]
            item["lon"] = location["GIS_Longitude"]

            if location["Province"] in ["EC", "NC"]:
                item["lat"], item["lon"] = item["lon"], item["lat"]

            if location.get("Town_City") is not None:
                item["city"] = location.get("Town_City").title().replace("'S", "'s")

            item["state"] = ZA_PROVINCES.get(location["Province"])

            if location.get("StreetAddress") is not None:
                item["addr_full"] = (
                    clean_address(location.get("StreetAddress")).title().replace("'S", "'s").replace("`S", "'s")
                )

            if location.get("PostalAddress") is not None:
                item["extras"]["addr:postal"] = (
                    clean_address(location.get("PostalAddress")).title().replace("'S", "'s").replace("`S", "'s")
                )

            if match := re.match(r"^.*\s(\d\d\d\d)$", item["addr_full"]):
                item["postcode"] = match[1]

            item["name"] = (
                location.get("Official_Institution_Name")
                .title()
                .replace("'S", "'s")
                .replace("`S", "'s")
                .replace(" P/S", " Primary School")
                .replace(" S/S", " Secondary School")
                .replace(" C/S", " Combined School")
                .replace(" I/S", " Intermediate School")
            )

            item["phone"] = location["Telephone"]

            apply_yes_no(Extras.FEE, item, location["NoFeeSchool"] == "NO", False)

            if location["Sector"].lower() == "public":
                item.update(ZA_EDUCATION_DEPARTMENTS.get(location["Province"]))
                item["extras"]["operator:type"] = "government"
            elif location["Sector"].lower() == "independent":
                item["extras"]["operator:type"] = "private"
            else:
                self.crawler.stats.inc_value(f"atp/{self.name}/unknown_operator_tyoe/{location['Sector']}")

            if location["Phase_PED"] == "PRIMARY SCHOOL":
                item["extras"]["school"] = "primary"
                apply_category(Categories.SCHOOL, item)
            elif location["Phase_PED"] == "SECONDARY SCHOOL":
                item["extras"]["school"] = "secondary"
                apply_category(Categories.SCHOOL, item)
            elif location["Phase_PED"] == "COMBINED SCHOOL":
                item["extras"]["school"] = "primary;secondary"
                apply_category(Categories.SCHOOL, item)
            elif location["Phase_PED"] == "SPECIAL NEEDS EDUCATION":
                item["extras"]["school"] = "special_education_needs"
                apply_category(Categories.SCHOOL, item)
            elif location["Phase_PED"] == "INTERMEDIATE SCHOOL":
                apply_category(Categories.SCHOOL, item)
            elif location["Phase_PED"] == "ECD":
                apply_category(Categories.KINDERGARTEN, item)
            else:
                self.crawler.stats.inc_value(f"atp/{self.name}/unknown_phase/{location['Phase_PED']}")

            if location["Type_DoE"] == "SPECIAL NEEDS EDUCATION CENTRE":
                item["extras"]["school"] = "special_education_needs"
                apply_category(Categories.SCHOOL, item)

            yield item
